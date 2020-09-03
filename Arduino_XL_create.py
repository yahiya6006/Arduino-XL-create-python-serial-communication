import serial
#import msxlt
from tkinter import *
import sys
import openpyxl as xl

class XLC:
    def create(self, data, name, sheetname):
        """This Method is going to create an XL sheet with first input as a sting
            i,e "Time,Data" etc which will be coloums and the secand input as the
            Filename
            [INFO] file extention is set to .xsml as a default
            and the third input as the name of the sheet"""
        # Importing Font to change the font 
        from openpyxl.styles import Font
        self.data = data
        self.name = name + ".xlsx"
        self.sheetname = sheetname

        # Created a tupple that contains all the names of coloums in an XL sheet
        self.colum = ('A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R'
                      ,'S','T','U','V')

        # Checking if file is already available
        self.cf = XLC.checkfile(self,self.name)
        #print(self.cf)
        if self.cf == True:
            print("[INFO] File {} already exists".format(self.name))
            userAns = input("[INFO] Do you want to over write the data \n[INFO] over writing the data will erase all the data stored in {} y/n: ".format(self.filename))
            if userAns == "y" or userAns == "Y":
                wb = xl.Workbook()
                wb.sheetnames
                sheet = wb.active
                sheet.title = self.sheetname

                # Seperating the received data using split
                self.data = self.data.split(",")
                for i in range(0, len(self.data)):

                    # When for loop starts we select 1st coloum i.e A1 simillarly B1 C1....
                    # using sheet[self.colum[i]+'1'] and the data is stored in it
                    sheet[self.colum[i]+'1'] = self.data[i]

                    # Setting the Font to Times New Roman 
                    sheet[self.colum[i]+'1'].font = Font(name='Times New Roman', bold=True, size=14)

                    # Setting the coloum width
                    sheet.column_dimensions[self.colum[i]].width = len(self.data[i])+12
                wb.save(self.name)
                print("{} saved".format(self.filename))
                return

            if userAns == "n" or userAns == "N":
                return
            
        if self.cf == False:
            # Creating the workbook
            wb = xl.Workbook()
            wb.sheetnames
            sheet = wb.active
            sheet.title = self.sheetname

            # Seperating the received data using split
            self.data = self.data.split(",")
            for i in range(0, len(self.data)):

                # When for loop starts we select 1st coloum i.e A1 simillarly B1 C1....
                # using sheet[self.colum[i]+'1'] and the data is stored in it
                sheet[self.colum[i]+'1'] = self.data[i]

                # Setting the Font to Times New Roman 
                sheet[self.colum[i]+'1'].font = Font(name='Times New Roman', bold=True, size=14)

                # Setting the coloum width
                sheet.column_dimensions[self.colum[i]].width = len(self.data[i])+12
            wb.save(self.name)
            print("{} saved".format(self.filename))
            return

    def checkfile(self, filename):
        import os
        self.filename = filename
        self.available = None
        self.BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        #print(self.BASE_DIR)
        with os.scandir(self.BASE_DIR) as entrys:
            for entry in entrys:
                #print(entry.name)
                if entry.name == self.filename:
                    self.available = True
                
            if self.available == True:
                return True
            if self.available != True:
                return False
        

    def updateData(self, data, filename):
        """This Method will update the values in the excel sheet where
            data shouled be a combined string seperated by (,)
            ex: "Time,date,status" etc where Time Date status are different data

            Filename is the name of the file where data needs to be updated
            [INFO] File extention is set to .xsml format"""
        self.data = data
        #self.filename = filename + ".xlsx"
        self.index = "index.txt"
        self.colum = ('A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R'
                      ,'S','T','U','V')
        self.cf = XLC.checkfile(self, self.index )
        #print(self.cf)
        if self.cf == True:
            self.resultFile = open(self.index, 'r')
            self.row = int(self.resultFile.read())
            #print(self.row)
            self.resultFile.close()
        if self.cf == False:
            print("[WARNING] index.txt file is missing. Neglect this message if this program is running 1st time on your system")
            self.row = 2
        
        # Looding the document
        self.filename = filename + ".xlsx"
        wb = xl.load_workbook(self.filename)
        sheet = wb.sheetnames
        sheet = wb.active

        if sheet['A2'].value == None:
            self.row = 2

        self.data = self.data.split(",")
        for i in range(0, len(self.data)):
            sheet[self.colum[i]+str(self.row)] = self.data[i]
            #print(self.colum[i]+str(self.row))

        self.row = self.row + 1
        self.resultFile = open('index.txt', 'w')
        self.resultFile.write(str(self.row))
        self.resultFile.close()
        wb.save(self.filename)
        return

    def time_date(self):
        """This method will give you current Time and Date"""
        from datetime import date
        from datetime import datetime

        self.today = date.today() # Output is: 2020-05-19
        # Need to change that in a format 19/05/2020
        self.today_format = self.today.strftime("%d/%m/%Y") #Output is: 19/05/2020

        self.now = datetime.now()
        self.current_time = self.now.strftime("%I:%M:%S %p") # %I- Hour 12hr foemat %M- minitue %S- sec %p- AM/PM
        # Both self.current_time and self.today_format are in string format

        return self.current_time, self.today_format

class Config:
    com_port = None
    bard_rate = None
    def __init__(self, root):
        self.root = root
        #------------- Setting Window Up ----------------------------
        self.root.minsize(320,180)
        self.root.maxsize(320,180)
        self.root.geometry('300x200+500+110')
        self.root.title('Arduino Serial')
        self.root.config(bg = '#292929')
        #-------------- Adding label, button and entry widget --------------
        cmp_frame = Frame(self.root)
        cmp_frame.pack(side=TOP, fill=X)
        cmp_frame.config(bg='#292929')
        Label(cmp_frame, text='YahiTech', bg='#292929', fg='#D9D9D9',
              font=('times new roman','20','bold')).pack(side=LEFT, padx=25, pady=10)
        #-------------- COM PORT -----------------
        comPort_frame = Frame(self.root)
        comPort_frame.pack(side=TOP, fill=X)
        comPort_frame.config(bg='#292929')
        Label(comPort_frame, text='COM PORT ', bg='#292929', fg='#D9D9D9',
              font=('times new roman','14','')).pack(side=LEFT, padx=15)
        self.comPort_entry = Entry(comPort_frame, font=('times new roman','12',''), bg='#292929',
                           fg='#ffffff', insertbackground='white', bd=2)
        self.comPort_entry.pack(side=LEFT)
        #-------------- BARD RATE -----------------
        bardRate_frame = Frame(self.root)
        bardRate_frame.pack(side=TOP, fill=X, pady=7)
        bardRate_frame.config(bg='#292929')
        Label(bardRate_frame, text='BARD RATE', bg='#292929', fg='#D9D9D9',
              font=('times new roman','14','')).pack(side=LEFT, padx=15)
        self.bardRate_entry = Entry(bardRate_frame, font=('times new roman','12',''), bg='#292929',
                           fg='#ffffff', insertbackground='white', bd=2)
        self.bardRate_entry.pack(side=LEFT)
        #-------------- Button -----------------
        self.Btn_frame =Frame(self.root)
        self.Btn_frame.pack(side=TOP, fill=BOTH)
        self.Btn_frame.config(bg='#292929')
        lgn_btn = Button(self.Btn_frame, text='OK', relief=RAISED, bd=2, bg='#595959',fg='#ffffff',
                         font = ('times new roman','14','bold'), command=self.get_OK_info)
        #lgn_btn.focus_set()
        #lgn_btn.bind("<Return>", lambda event: self.get_lgn_info())
        lgn_btn.pack(side=RIGHT, padx=40, pady=5)

    def get_OK_info(self):
        Config.com_port = self.comPort_entry.get()
        Config.bard_rate = self.bardRate_entry.get()
        self.root.destroy()

    def get_info():
        return com_port, bard_rate


root = Tk()
cf = Config(root)
root.mainloop()
comport, bardrate = str(Config.com_port), int(Config.bard_rate)
print(comport.upper(), bardrate)
try:
    port = serial.Serial(comport.upper(), bardrate)
except serial.serialutil.SerialException:
    print('[INFO] Arduino is not connected.')
    sys.exit()
XL = XLC()
XL.create("Date,Time,Count","Object_counting","Yahitech")
while(port.isOpen()):
    data = str(port.readline())
    data = 111
    time, date = XL.time_date()
    data = data[2:len(data)-3]
    try:
        XL.updateData(date+','+time+','+data,"Object_counting")
    except:
        pass
    print(data)

    

